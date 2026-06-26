# -*- coding: utf-8 -*-
"""
core/table_ops.py — 工作区表格读写（table_ops 工具的核心逻辑）

能力（仅工作区内文件，不做复杂数据分析——交给 calculator 配合）：
  read   读 .xlsx → 返回 markdown 表格（供 Agent 理解数据）
  write  从 markdown 表格 / CSV → 生成 .xlsx

数据分析（求和/筛选/统计）不在此工具范围，由 calculator 工具配合完成：
  Agent 用 table_read 读出数据 → 用 calculator 计算。
"""
import os
import csv
import io
import re
import logging

log = logging.getLogger(__name__)


def read_xlsx(path: str, max_rows: int = 50) -> dict:
    """读取 xlsx，返回 markdown 表格 + 维度信息。

    Args:
        path: xlsx 文件绝对路径
        max_rows: 最多读取行数（防止超大表爆 token）

    Returns:
        {"ok": bool, "markdown": str, "rows": int, "cols": int, "sheets": list, "error": str}
    """
    if not os.path.exists(path):
        return {"ok": False, "error": "文件不存在: %s" % os.path.basename(path)}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        # 读第一个 sheet（多 sheet 场景让 Agent 自己决定要不要切）
        ws = wb[sheets[0]]
        rows_data = []
        max_col = 0
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri >= max_rows:
                break
            # 转字符串，None → 空串
            str_row = [("" if c is None else str(c)) for c in row]
            max_col = max(max_col, len(str_row))
            rows_data.append(str_row)
        wb.close()

        if not rows_data:
            return {"ok": True, "markdown": "(空表格)", "rows": 0, "cols": 0, "sheets": sheets}

        # 补齐每行列数一致（openpyxl 可能返回不等长）
        for r in rows_data:
            while len(r) < max_col:
                r.append("")

        # 构建 markdown 表格（首行作为表头）
        lines = []
        for ri, r in enumerate(rows_data):
            lines.append("| " + " | ".join(r) + " |")
            if ri == 0:
                lines.append("| " + " | ".join(["---"] * max_col) + " |")
        markdown = "\n".join(lines)

        return {
            "ok": True,
            "markdown": markdown,
            "rows": len(rows_data),
            "cols": max_col,
            "sheets": sheets,
            "truncated": len(rows_data) >= max_rows,
        }
    except Exception as e:
        log.error("[TABLE] 读取 xlsx 失败: %s", str(e)[:150])
        return {"ok": False, "error": "读取失败: %s" % str(e)[:120]}


def write_xlsx(path: str, data: str, fmt: str = "markdown") -> dict:
    """从 markdown 表格或 CSV 文本生成 xlsx。

    Args:
        path: 目标 xlsx 绝对路径
        data: 表格数据文本（markdown 表格 或 CSV）
        fmt: "markdown" 或 "csv"

    Returns:
        {"ok": bool, "rows": int, "cols": int, "error": str}
    """
    try:
        from openpyxl import Workbook
        rows = _parse_table_data(data, fmt)
        if not rows:
            return {"ok": False, "error": "未解析出表格数据（检查格式：markdown 表格或 CSV）"}

        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for r in rows:
            ws.append(r)
        wb.save(path)

        cols = max(len(r) for r in rows) if rows else 0
        log.info("[TABLE] 写入 xlsx: %s (%d行×%d列)", os.path.basename(path), len(rows), cols)
        return {"ok": True, "rows": len(rows), "cols": cols, "name": os.path.basename(path)}
    except Exception as e:
        log.error("[TABLE] 写入 xlsx 失败: %s", str(e)[:150])
        return {"ok": False, "error": "写入失败: %s" % str(e)[:120]}


def _parse_table_data(data: str, fmt: str):
    """把文本解析成二维数组（行×列）。"""
    rows = []
    if fmt == "csv":
        reader = csv.reader(io.StringIO(data))
        for r in reader:
            rows.append(r)
    else:
        # markdown 表格：只取 | 分隔的行，跳过分隔线（|---|）
        for line in data.strip().split("\n"):
            line = line.strip()
            if not line or "|" not in line:
                continue
            # 跳过分隔行
            if re.match(r"^\|[\s\-:|]+\|$", line):
                continue
            # 拆分单元格
            cells = [c.strip() for c in line.split("|")]
            # 去掉首尾空（| a | b | 拆分后首尾是空串）
            if cells and cells[0] == "":
                cells = cells[1:]
            if cells and cells[-1] == "":
                cells = cells[:-1]
            if cells:
                rows.append(cells)
    return rows
