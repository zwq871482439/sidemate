# -*- coding: utf-8 -*-
"""
files/file_reader.py — 文件读取工具

从 routers/chat.py 提取的文件读取函数：
  - read_excel  — 读取 Excel 文件
  - read_doc    — 读取旧版 .doc 文件
  - read_csv    — 读取 CSV 文件

注意：函数名已去掉前缀下划线（公开 API）。
"""


def read_excel(file_path):
    """读取 Excel 文件"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            parts = []
            for sheet_name in wb.sheetnames[:5]:
                ws = wb[sheet_name]
                parts.append("== Sheet: %s ==" % sheet_name)
                row_count = 0
                for row in ws.iter_rows(max_row=50, values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        parts.append(" | ".join(cells))
                        row_count += 1
                if row_count == 50:
                    parts.append("... (更多行已省略)")
            return "\n".join(parts) if parts else "[Excel 文件为空]"
        finally:
            wb.close()
    except ImportError:
        try:
            import pandas as pd
            dfs = pd.read_excel(file_path, sheet_name=None, nrows=50)
            parts = []
            for name, df in list(dfs.items())[:5]:
                parts.append("== Sheet: %s ==" % name)
                parts.append(df.to_string(max_rows=50, max_cols=20))
            return "\n".join(parts) if parts else "[Excel 文件为空]"
        except ImportError:
            return "[Excel 读取失败: 需要 openpyxl 或 pandas 库]"
    except Exception as e:
        return "[Excel 读取失败: %s]" % str(e)[:100]


def read_doc(file_path):
    """读取旧版 .doc 文件"""
    import subprocess
    try:
        from docx import Document
        doc = Document(file_path)
        parts = [p.text for p in doc.paragraphs if p.text.strip()]
        if parts:
            return "\n".join(parts)
    except Exception:
        pass
    try:
        result = subprocess.run(["antiword", file_path], capture_output=True, text=True, timeout=10)
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout
    except (FileNotFoundError, Exception):
        pass
    try:
        import textract
        text = textract.process(file_path).decode("utf-8", errors="replace")
        if text.strip():
            return text
    except ImportError:
        pass
    except Exception:
        pass
    return "[.doc 文件读取失败: 建议将文件另存为 .docx 格式后重试]"


def read_csv(file_path):
    """读取 CSV 文件"""
    import csv
    parts = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= 100:
                    parts.append("... (更多行已省略)")
                    break
                parts.append(" | ".join(row))
    except Exception as e:
        return "[CSV 读取失败: %s]" % str(e)[:100]
    return "\n".join(parts) if parts else "[CSV 文件为空]"
